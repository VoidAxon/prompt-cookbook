#!/usr/bin/env python3
"""③転記用（パイプ区切り）テーブルを検査し、xlsx（無ければCSVへ降級）に変換する。

spec-impl-reconcile の出力 ③ を人手変換せずに Excel で開ける成果物へ落とすための再利用ツール。
- 入力: パイプ `|` 区切りのテキスト（1行目ヘッダ、以降1指摘=1行）。ファイル or stdin。
- 検査（機械チェック）: 全行のパイプ本数が揃うか / 先頭・末尾パイプが無いか / 非日本語スクリプト混入 /
  受控列の値が許可一覧に収まるか（既定で `分類`/`深刻度`/`対処区分` を照合。`--enum`/`--sections` で追加）。
- 出力: <out>.xlsx（openpyxl があれば）。無ければ <out>.csv（UTF-8 BOM・必要時クォート）へ降級し警告。

使い方:
    python pipe_to_xlsx.py <input_pipe.txt> <output_base>   # 明示: out\\Foo_issues （拡張子は付けない）
    python pipe_to_xlsx.py <input_pipe.txt> --name <対象名>  # 既定: %USERPROFILE%\\Documents\\<対象名>\\<対象名>_issues
    cat table.txt | python pipe_to_xlsx.py - --name <対象名>

受控列の照合:
    既定で `分類`(A1..E) / `深刻度`(高/中/低) / `対処区分`(新規抽出/横断新設/追補/修正/再抽出/確認) を常に照合する（列が無ければスキップ）。
    `--enum 列名=v1,v2,...` で任意列を追加・上書き（繰り返し可）。
    `--sections a,b,c` は `--enum 指摘箇所=a,b,c` の別名（指摘箇所は文書テンプレート依存＝渡した時のみ照合）。

出力先:
- <output_base> を明示した場合はそこへ書く。
- 省略し --name <対象名> を渡した場合、既定の %USERPROFILE%\\Documents\\<対象名>\\<対象名>_issues に書く（フォルダは自動作成）。
- 既存の同名ファイルは同フォルダの old\\ へ退避してから上書きする（黙って消さない）。

終了コード: 0=成功, 1=引数不足, 2=構造エラー（パイプ本数不一致・首尾パイプ）, 3=生成失敗。
非日本語スクリプト混入は警告のみ（μ 等の正当な記号もあるため、人が最終確認する前提）。
"""
import os
import sys
import re
import csv

# キリル/ギリシャ/ヘブライ/アラビア/タイ/ハングル等（生成タイポの検出用。日本語=かな/漢字は対象外）
FORBIDDEN = re.compile(r"[Ͱ-ϿЀ-ӿ֐-׿؀-ۿ฀-๿가-힯]")

# 本スキル自身の分類体系（領域・言語非依存の受控列）＝既定で常に照合する。
# 指摘箇所 は文書テンプレート依存のため既定に含めない（--sections / --enum で渡した時のみ照合）。
DEFAULT_ENUMS = {
    "分類": ["A1", "A2", "A3", "B", "C", "D", "E"],
    "深刻度": ["高", "中", "低"],
    "対処区分": ["新規抽出", "横断新設", "追補", "修正", "再抽出", "確認"],
}


def parse(text):
    """パイプ区切りテキストを行×セルの2次元リストへ。各セルはトリムする。"""
    rows = []
    for line in text.splitlines():
        if line.strip() == "":
            continue
        # 先頭・末尾パイプは構造エラーとして後段で検出するため、ここでは strip しない
        rows.append([c.strip() for c in line.split("|")])
    return rows


def check(rows, raw_lines, enums=None):
    """機械チェック。致命的エラーのリストと警告のリストを返す。

    enums: {ヘッダ名: 許可値リスト} の辞書。該当ヘッダの列を1行ずつ照合し、一覧外は警告。
           表にそのヘッダが無ければスキップ（列名で対応づけるので列位置に依存しない）。
    """
    errors, warnings = [], []
    if not rows:
        return ["入力が空です"], warnings
    ncol = len(rows[0])
    header = [h.strip() for h in rows[0]]
    col_idx = {h: j for j, h in enumerate(header)}  # ヘッダ名 → 列index
    enums = enums or {}
    for i, (cells, raw) in enumerate(zip(rows, raw_lines), 1):
        if len(cells) != ncol:
            errors.append(f"{i}行目: 列数 {len(cells)}（ヘッダは {ncol}）— パイプ本数不一致")
        if raw.startswith("|") or raw.rstrip().endswith("|"):
            errors.append(f"{i}行目: 先頭または末尾にパイプがあります")
        m = FORBIDDEN.search(raw)
        if m:
            warnings.append(f"{i}行目: 非日本語スクリプト文字 {m.group()!r} を検出（生成タイポの可能性・要確認）")
        if i == 1:
            continue  # ヘッダ行は値照合しない
        for name, allowed in enums.items():
            j = col_idx.get(name)
            if j is None or j >= len(cells):
                continue
            v = cells[j].strip()
            if v and v not in allowed:
                warnings.append(f"{i}行目: {name} {v!r} が許可値一覧に無い {allowed}（枠外値/かっこ書き残り/§付き/複数併記/生成タイポを確認）")
    return errors, warnings


def write_xlsx(rows, path):
    """openpyxl で xlsx を書く。ヘッダ太字・オートフィルタ・折返し・列幅調整。失敗時は例外。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "issues"
    for r in rows:
        ws.append(r)
    for c in range(1, len(rows[0]) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(rows[0]))}{len(rows)}"
    # 列幅: 見出し名で判定。長文列（指摘内容/備考）は広め＋折返し
    wide = {"指摘内容", "備考"}
    for idx, name in enumerate(rows[0], 1):
        col = get_column_letter(idx)
        ws.column_dimensions[col].width = 80 if name in wide else min(max(len(name) + 4, 10), 24)
        if name in wide:
            for row in range(2, len(rows) + 1):
                ws.cell(row=row, column=idx).alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(path)


def write_csv(rows, path):
    """UTF-8 BOM・必要時クォートの CSV（Excel が直接開ける降級版）。"""
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        csv.writer(f, quoting=csv.QUOTE_MINIMAL).writerows(rows)


def parse_opts(argv):
    """argv[2:] を解析して (out_base, user_enums) を返す。

    --name <対象名>       → 既定 %USERPROFILE%\\Documents\\<対象名>\\<対象名>_issues（フォルダ作成）
    --sections a,b,c      → 指摘箇所の許可値（= --enum 指摘箇所=a,b,c の別名）
    --enum 列名=v1,v2,... → 任意の受控列の許可値（繰り返し可・既定を上書き/追加）
    それ以外の位置引数     → 明示 output_base
    """
    args = argv[2:]
    name = None
    user_enums = {}
    positional = None
    i = 0
    while i < len(args):
        a = args[i]
        if a == "--name" and i + 1 < len(args):
            name = args[i + 1]; i += 2; continue
        if a == "--sections" and i + 1 < len(args):
            user_enums["指摘箇所"] = [s.strip() for s in args[i + 1].split(",") if s.strip()]; i += 2; continue
        if a == "--enum" and i + 1 < len(args):
            spec = args[i + 1]; i += 2
            if "=" in spec:
                col, vals = spec.split("=", 1)
                user_enums[col.strip()] = [s.strip() for s in vals.split(",") if s.strip()]
            continue
        if not a.startswith("--"):
            positional = a
        i += 1

    if positional:
        out_base = positional
    elif name:
        home = os.environ.get("USERPROFILE") or os.path.expanduser("~")
        d = os.path.join(home, "Documents", name)
        os.makedirs(d, exist_ok=True)
        out_base = os.path.join(d, name + "_issues")
    else:
        out_base = None
    return out_base, user_enums


def stash_old(path):
    """既存の同名ファイルを同フォルダの old\\ へ退避（衝突時は連番）。黙って上書きしない。"""
    if not os.path.exists(path):
        return
    d = os.path.dirname(path) or "."
    base = os.path.basename(path)
    oldd = os.path.join(d, "old")
    os.makedirs(oldd, exist_ok=True)
    dest = os.path.join(oldd, base)
    root, ext = os.path.splitext(base)
    n = 1
    while os.path.exists(dest):
        dest = os.path.join(oldd, f"{root}_{n}{ext}")
        n += 1
    os.replace(path, dest)


def main(argv):
    if len(argv) < 3:
        print(__doc__)
        return 1
    src = argv[1]
    out_base, user_enums = parse_opts(argv)
    if not out_base:
        print("[ERROR] 出力先が未指定です。<output_base> か --name <対象名> を渡してください。", file=sys.stderr)
        print(__doc__)
        return 1
    text = sys.stdin.read() if src == "-" else open(src, encoding="utf-8").read()
    raw_lines = [l for l in text.splitlines() if l.strip() != ""]
    rows = parse(text)

    enums = dict(DEFAULT_ENUMS)
    enums.update(user_enums)  # --enum / --sections が既定を上書き・追加
    errors, warnings = check(rows, raw_lines, enums)
    for w in warnings:
        print(f"[WARN] {w}", file=sys.stderr)
    if errors:
        for e in errors:
            print(f"[ERROR] {e}", file=sys.stderr)
        print("[NG] 機械チェックに失敗。③を修正してから再実行してください。", file=sys.stderr)
        return 2

    try:
        path = out_base + ".xlsx"
        stash_old(path)
        write_xlsx(rows, path)
        print(f"[OK] xlsx を生成: {path}")
        return 0
    except ImportError:
        path = out_base + ".csv"
        stash_old(path)
        write_csv(rows, path)
        print(f"[WARN] openpyxl が無いため xlsx を生成できず CSV へ降級: {path}", file=sys.stderr)
        print("       真の xlsx が必要なら: pip install openpyxl（または PowerShell の ImportExcel）", file=sys.stderr)
        return 0
    except Exception as e:  # 生成失敗も CSV へ降級して成果物は必ず残す
        path = out_base + ".csv"
        try:
            stash_old(path)
        except Exception:
            pass
        write_csv(rows, path)
        print(f"[WARN] xlsx 生成に失敗（{e}）。CSV へ降級: {path}", file=sys.stderr)
        if isinstance(e, PermissionError) or "WinError 32" in str(e):
            print("       → 出力先 xlsx が使用中（Excel で開いていないか）。閉じてから再実行してください。", file=sys.stderr)
        return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
