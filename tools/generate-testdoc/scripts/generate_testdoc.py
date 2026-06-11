#!/usr/bin/env python3
"""Generate test specification Excel document from JSON input."""

import argparse
import glob
import json
import re
import sys
import zipfile
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    print("[ERROR] openpyxl が必要です。pip install openpyxl でインストールしてください。")
    sys.exit(1)

SKILL_ROOT = Path(__file__).parent.parent
TEMPLATE_PATH = SKILL_ROOT / "templates" / "test_spec_template.xlsx"

# テスト観点シート ヘッダー位置（PR 由来の場合のみ書き込む）
VP_HEADER_RECORD = (3, 3)   # C3: kintoneレコード（PR タイトル、cybozu リンク付き）
VP_HEADER_PR = (4, 3)       # C4: #<PR番号>（PR リンク付き）

# ハイパーリンクの見た目（青字・下線）
HYPERLINK_FONT = Font(color="FF0563C1", underline="single")

# テスト観点シート 列位置
VP_DATA_START = 6
VP_COL_NO = 2       # B: No.
VP_COL_GROUP = 3    # C: テスト観点（グループ名）
VP_COL_ITEMS = 4    # D: 確認項目（箇条書き）

# テストケースシート 列位置
TC_DATA_START = 9
TC_COL_NO = 1       # A: 項番
TC_COL_NAME = 2     # B: テスト観点（グループ行=グループ名、詳細行=テスト名）
TC_COL_PRECOND = 3  # C: 設定・前提条件
TC_COL_STEPS = 4    # D: 入力値・操作手順
TC_COL_EXPECTED = 5 # E: 期待値
TC_COL_RESULT = 6   # F: 結果
TC_COL_LAST = 8     # H: 実施日（グループ行の塗りつぶし終端）

# グループ行の背景色（テスト進捗率セルと同色）
GROUP_ROW_FILL = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")


def load_json(input_path=None):
    if input_path:
        with open(input_path, encoding="utf-8") as f:
            return json.load(f)
    return json.load(sys.stdin)


def resolve_output_path(output_dir, kintone_id, title):
    output_dir = Path(output_dir).expanduser()
    output_dir.mkdir(parents=True, exist_ok=True)

    existing = glob.glob(str(output_dir / f"【{kintone_id}】*.xlsx"))
    if len(existing) > 1:
        print(f"[WARNING] 同一KintoneIDのファイルが複数存在します。最初のファイルを更新します: {existing[0]}")
    if existing:
        return Path(existing[0])
    return output_dir / f"【{kintone_id}】 {title}.xlsx"


def load_workbook(output_path):
    if output_path.exists():
        return openpyxl.load_workbook(output_path)
    if TEMPLATE_PATH.exists():
        return openpyxl.load_workbook(TEMPLATE_PATH)
    wb = openpyxl.Workbook()
    wb.active.title = "テスト観点"
    wb.create_sheet("テストケース")
    wb.create_sheet("エビデンス")
    return wb


def find_sheet(wb, keywords):
    for name in wb.sheetnames:
        if any(kw in name for kw in keywords):
            return wb[name]
    return None


def ensure_evidence_sheet(wb):
    """エビデンスシートがなければ追加する。"""
    if not any("エビデンス" in name for name in wb.sheetnames):
        wb.create_sheet("エビデンス")


def set_text_cell(ws, row, col, value):
    """値をテキスト形式で書き込む。number_format="@" により Excel が数式と誤認識しない。"""
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = "@"
    cell.alignment = Alignment(wrap_text=True, vertical="top")


def write_pr_header(ws, pr_title, record_url, pr_no, pr_url):
    """PR 由来の場合、テスト観点シートのヘッダーに PR 情報を書き込む。

    C3: PR タイトル（KintoneID 除去済み）。record_url（cybozu リンク）があればリンク化。
    C4: #<PR番号>。pr_url があればリンク化。
    """
    if pr_title:
        row, col = VP_HEADER_RECORD
        cell = ws.cell(row=row, column=col, value=pr_title)
        cell.number_format = "@"
        if record_url:
            cell.hyperlink = record_url
            cell.font = HYPERLINK_FONT
    if pr_no:
        row, col = VP_HEADER_PR
        cell = ws.cell(row=row, column=col, value=f"#{pr_no}")
        cell.number_format = "@"
        if pr_url:
            cell.hyperlink = pr_url
            cell.font = HYPERLINK_FONT


def write_test_viewpoints(ws, viewpoints):
    """テスト観点シートにグループ単位でデータを書き込む。"""
    for i, vp in enumerate(viewpoints):
        row = VP_DATA_START + i
        set_text_cell(ws, row, VP_COL_NO, i + 1)
        set_text_cell(ws, row, VP_COL_GROUP, vp["group"])
        items_text = "\n".join(f"・{item}" for item in vp["items"])
        set_text_cell(ws, row, VP_COL_ITEMS, items_text)


def write_test_cases(ws, test_viewpoints, test_cases):
    """テストケースシートにグループ行・詳細テストケース行を書き込む。

    テスト観点シートのグループ順に合わせてグループ行を挿入し、
    その配下に詳細テストケースを並べる。
    """
    groups_order = [vp["group"] for vp in test_viewpoints]
    cases_by_group: dict[str, list] = {}
    for tc in test_cases:
        g = tc.get("group", "")
        if g not in cases_by_group:
            cases_by_group[g] = []
        cases_by_group[g].append(tc)

    current_row = TC_DATA_START
    item_no = 1

    for group in groups_order:
        cases = cases_by_group.get(group, [])
        if not cases:
            continue

        # グループ行（結果=対象外、項番なし、背景色あり）
        ws.cell(row=current_row, column=TC_COL_NO).value = None  # テンプレート数式を上書きしてクリア
        set_text_cell(ws, current_row, TC_COL_NAME, group)
        set_text_cell(ws, current_row, TC_COL_RESULT, "対象外")
        for col in range(TC_COL_NO, TC_COL_LAST + 1):
            ws.cell(row=current_row, column=col).fill = GROUP_ROW_FILL
        current_row += 1

        # 詳細テストケース行
        for tc in cases:
            steps_text = "\n".join(f"{j + 1}. {step}" for j, step in enumerate(tc["steps"]))
            set_text_cell(ws, current_row, TC_COL_NO, item_no)
            set_text_cell(ws, current_row, TC_COL_NAME, tc["name"])
            set_text_cell(ws, current_row, TC_COL_PRECOND, tc["precondition"])
            set_text_cell(ws, current_row, TC_COL_STEPS, steps_text)
            set_text_cell(ws, current_row, TC_COL_EXPECTED, tc["expected"])
            current_row += 1
            item_no += 1


def fix_inline_strings(xlsx_path):
    """openpyxlが生成する inlineStr セルを shared strings に変換して Excel の #NAME? を回避する。"""
    xlsx_path = Path(xlsx_path)

    with zipfile.ZipFile(xlsx_path, "r") as zin:
        files = {name: zin.read(name) for name in zin.namelist()}

    sheet_names = [n for n in files if n.startswith("xl/worksheets/") and n.endswith(".xml")]

    def _decode_xml(text):
        """&#NNNN; などの XML 文字参照を Unicode 文字に戻す。"""
        text = re.sub(r"&#(\d+);", lambda m: chr(int(m.group(1))), text)
        text = re.sub(r"&#x([0-9a-fA-F]+);", lambda m: chr(int(m.group(1), 16)), text)
        return text.replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">").replace("&quot;", '"')

    def _escape_xml(text):
        return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    # inlineStr の生テキスト（XML エンティティ含む）→ shared strings インデックス
    encoded_to_idx = {}
    decoded_list = []

    for name in sheet_names:
        for encoded in re.findall(
            r't="inlineStr"[^>]*><is><t[^>]*>(.*?)</t></is></c>',
            files[name].decode("utf-8"),
            re.DOTALL,
        ):
            if encoded not in encoded_to_idx:
                encoded_to_idx[encoded] = len(decoded_list)
                decoded_list.append(_decode_xml(encoded))

    if not decoded_list:
        return

    # 各シートの inlineStr セルを t="s" に置換し、行高さをExcelに委ねる
    for name in sheet_names:
        xml = files[name].decode("utf-8")

        def _replace(m):
            attrs = re.sub(r'\s*t="inlineStr"', "", m.group(1))
            idx = encoded_to_idx.get(m.group(2))
            return f'<c{attrs} t="s"><v>{idx}</v></c>' if idx is not None else m.group(0)

        xml = re.sub(
            r'<c([^>]+)t="inlineStr"[^>]*><is><t[^>]*>(.*?)</t></is></c>',
            _replace,
            xml,
            flags=re.DOTALL,
        )
        # ht / customHeight を除去 → Excel がファイルを開いた際に自動で行高さを計算する
        xml = re.sub(r'\s+ht="[^"]*"', "", xml)
        xml = re.sub(r'\s+customHeight="[^"]*"', "", xml)
        files[name] = xml.encode("utf-8")

    # sharedStrings.xml を生成（改行・先頭末尾空白がある場合は xml:space="preserve"）
    count = len(decoded_list)
    ss_items = []
    for text in decoded_list:
        escaped = _escape_xml(text)
        preserve = ' xml:space="preserve"' if ("\n" in text or text != text.strip()) else ""
        ss_items.append(f"<si><t{preserve}>{escaped}</t></si>")

    files["xl/sharedStrings.xml"] = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
        f' count="{count}" uniqueCount="{count}">'
        + "".join(ss_items)
        + "</sst>"
    ).encode("utf-8")

    # [Content_Types].xml に sharedStrings エントリを追加
    ct = files["[Content_Types].xml"].decode("utf-8")
    if "sharedStrings" not in ct:
        ct = ct.replace(
            "</Types>",
            '<Override PartName="/xl/sharedStrings.xml"'
            ' ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
            "</Types>",
        )
        files["[Content_Types].xml"] = ct.encode("utf-8")

    # workbook.xml.rels に sharedStrings リレーションを追加
    rels_name = "xl/_rels/workbook.xml.rels"
    if rels_name in files:
        rels = files[rels_name].decode("utf-8")
        if "sharedStrings" not in rels:
            rids = re.findall(r'Id="rId(\d+)"', rels)
            next_id = max(int(r) for r in rids) + 1 if rids else 10
            rels = rels.replace(
                "</Relationships>",
                f'<Relationship Id="rId{next_id}"'
                f' Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings"'
                f' Target="sharedStrings.xml"/></Relationships>',
            )
            files[rels_name] = rels.encode("utf-8")

    # 上書き保存
    tmp = xlsx_path.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in files.items():
            zout.writestr(name, data)
    xlsx_path.unlink()
    tmp.rename(xlsx_path)


def main():
    parser = argparse.ArgumentParser(description="テスト仕様書Excelを生成する")
    parser.add_argument("--kintone-id", required=True, help="KintoneチケットID")
    parser.add_argument("--title", required=True, help="変更タイトル")
    parser.add_argument("--input", help="JSONファイルパス（省略時はstdin）")
    parser.add_argument("--output", default="~/Documents/docs/testcase/", help="出力ディレクトリ")
    parser.add_argument("--pr-no", help="PR番号。指定時にテスト観点シートのC3/C4へPR情報を書き込む")
    parser.add_argument("--pr-url", help="PRのURL（C4のリンク先）")
    parser.add_argument("--pr-title", help="C3に書くPRタイトル（KintoneID除去済み。省略時は--title）")
    parser.add_argument("--record-url", help="PR本文から抽出したcybozuリンク（C3のリンク先、省略可）")
    args = parser.parse_args()

    try:
        data = load_json(args.input)
    except (json.JSONDecodeError, FileNotFoundError) as e:
        print(f"[ERROR] JSON読み込みに失敗しました: {e}")
        sys.exit(1)

    output_path = resolve_output_path(args.output, args.kintone_id, args.title)
    wb = load_workbook(output_path)

    sheet_viewpoints = find_sheet(wb, ["観点"])
    sheet_cases = find_sheet(wb, ["ケース"])

    if sheet_viewpoints is None or sheet_cases is None:
        print(f"[ERROR] テンプレートに 'テスト観点' または 'テストケース' シートが見つかりません。")
        print(f"  シート一覧: {wb.sheetnames}")
        sys.exit(1)

    ensure_evidence_sheet(wb)
    if args.pr_no:
        write_pr_header(
            sheet_viewpoints,
            args.pr_title or args.title,
            args.record_url,
            args.pr_no,
            args.pr_url,
        )
    write_test_viewpoints(sheet_viewpoints, data["test_viewpoints"])
    write_test_cases(sheet_cases, data["test_viewpoints"], data["test_cases"])

    wb.save(output_path)
    fix_inline_strings(output_path)

    total_cases = len(data["test_cases"])
    print(f"[完了] テスト仕様書を出力しました。")
    print(f"  ファイル: {output_path}")
    print(f"  テスト観点グループ: {len(data['test_viewpoints'])} グループ")
    print(f"  テストケース: {total_cases} 件")


if __name__ == "__main__":
    main()
